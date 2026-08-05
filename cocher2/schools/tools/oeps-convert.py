"""Opera Estate Primary 2026 Term 1 staff timetable (.xlsx) → canonical entries.

The layout is TRANSPOSED against Park View's: days run down the rows, periods
run across the columns, and a merged range — not a repeated value — is how a
double period is written. So merges are expanded first; without that, openpyxl
hands back the value only in a range's top-left cell and every double period
would silently become a single.

Each day is a three-row block:
    row A     subject codes, one per period column
    row A+1   the DAY name in column 1, then class codes
    row A+2   an UNTIMED duty post in column 1 (PEDESTRIAN, HALL, TEAM LEADER…)

Blocks are located by the day name, not by row number, because one sheet's
header sits a row lower than the other eighty.
"""
import openpyxl, json, re, sys, collections

SRC = sys.argv[1]
DAYS = {'MON': 'Mon', 'TUE': 'Tue', 'WED': 'Wed', 'THU': 'Thu', 'FRI': 'Fri'}
ORDER = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']

# A class code in this school: 1C … 6H, plus 6FDN (the P6 Foundation class).
CLASS_RE = re.compile(r'^[1-6](?:[A-Z]{1,3}|FDN)\d?$')
# Duty posts, written into the grid rather than into the duty column. Some wrap
# over the two rows of a block ("BICYCLE" above "BAY"), so they are matched
# after the two halves are joined.
DUTY = {'CANTEEN', 'FIELD', 'HALL', 'GREEN CORRIDOR', 'BUS BAY', 'BUS STOP',
        'BUS STOP GATE', 'BICYCLE BAY', 'PEDESTRIAN CROSSING', 'PED CROSSING',
        'GATE', 'CORRIDOR', 'LATE REPORTING', 'P1 TRANSIT', 'PARADE SQUARE'}


def cellgrid(ws):
    """Every cell, with merged ranges filled out so a span reads on all its columns."""
    g = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            g[(r, c)] = '' if v is None else re.sub(r'\s+', ' ', str(v).strip())
    for m in ws.merged_cells.ranges:
        v = g.get((m.min_row, m.min_col), '')
        for r in range(m.min_row, m.max_row + 1):
            for c in range(m.min_col, m.max_col + 1):
                g[(r, c)] = v
    return g


# The same post is written several ways across 81 hand-maintained sheets. Left
# alone, one duty reads as three different places.
ALIAS = {'PED CROSSING': 'PEDESTRIAN CROSSING', 'GREEN': 'GREEN CORRIDOR',
         'PARADE SQ': 'PARADE SQUARE'}


def tidy(s):
    s = re.sub(r'\s+', ' ', (s or '').strip())
    return ALIAS.get(s.upper(), s)


def resolve(subj, klass, codes):
    """One period column → (title, class, kind).

    Four shapes appear in this file and each is a different answer:
      MA / 5D           a lesson
      5D / MA           the same lesson, written the other way up
      CANTEEN / ''      a duty post
      BICYCLE / BAY     a duty post whose label wrapped over both rows
    """
    s, k = tidy(subj), tidy(klass)
    if not (s or k):
        return None

    # Written upside down: the class landed in the subject row.
    if s and k and CLASS_RE.match(s.upper()) and not CLASS_RE.match(k.upper()) \
            and k.upper() in codes:
        s, k = k, s

    su, ku = s.upper(), k.upper()
    if su and su == ku:                       # the same label repeated in both rows
        k = ''
        ku = ''

    # A duty whose label wrapped: "GREEN" over "CORRIDOR", "BICYCLE" over "BAY".
    joined = f'{su} {ku}'.strip()
    if joined in DUTY:
        return (joined.title(), None, 'duty')
    if su in DUTY:
        # "FIELD" over "FUN FRI" — the post, with what is happening there.
        return (su.title(), k or None, 'duty')
    if not s and ku in DUTY:
        return (ku.title(), None, 'duty')

    if s and su in codes:                     # a subject the legend knows
        return (s, k or None, kind_of(su))
    if not s:                                 # the label lives in the class row alone
        return (k, None, kind_of(ku))
    return (s, k or None, kind_of(su))


def kind_of(code):
    c = (code or '').upper()
    if c in DUTY:
        return 'duty'
    if re.search(r'^ASM$|ASSEMBLY|FTGP|RECESS|BLOCKED|FUN FRI|CCE|COUNSELLING|TRANSIT', c):
        return 'school'
    return 'lesson'


wb = openpyxl.load_workbook(SRC, data_only=True)
teachers, legend_subj, legend_room, tag, odd = [], {}, {}, '', collections.Counter()

for name in wb.sheetnames:
    ws = wb[name]
    g = cellgrid(ws)
    tag = tag or g.get((1, 1), '')

    hdr = next((r for r in range(1, ws.max_row + 1)
                if g.get((r, 1), '').upper().startswith('DAY/PERIOD')), None)
    if hdr is None:
        continue

    # "3 0830-0900" → the real clock times this column stands for.
    slots = {}
    for c in range(2, ws.max_column + 1):
        m = re.match(r'^\s*(\d+)\s+(\d{4})-(\d{4})\s*$', g.get((hdr, c), ''))
        if m:
            slots[c] = (f'{m.group(2)[:2]}:{m.group(2)[2:]}', f'{m.group(3)[:2]}:{m.group(3)[2:]}')
    if not slots:
        continue

    dayrows = [(r, DAYS[g[(r, 1)].upper()]) for r in range(hdr + 1, ws.max_row + 1)
               if g.get((r, 1), '').upper() in DAYS]
    stop = max(r for r, _ in dayrows) + 3

    # Legend: short code → full name, for subjects and for rooms. First writing
    # wins — one sheet glosses LSP1 with a room number, and a later overwrite
    # would leave "LSP for P1" reading as "C2-03B" for every school.
    for r in range(stop, ws.max_row + 1):
        for c in range(2, ws.max_column):
            k, v = g.get((r, c), ''), g.get((r, c + 1), '')
            if not (k and v) or k.upper() in ('SUBJECTS', 'RESOURCES') or k == v:
                continue
            room = re.match(r'^[A-Z]\d-|LAB|ROOM|RES$', k.upper()) or re.match(r'^[A-Z]\d-\d', v.upper())
            (legend_room if room else legend_subj).setdefault(k, v)

    codes = {k.upper() for k in legend_subj} | {
        'EL', 'MA', 'MT', 'SCI', 'SS', 'PE', 'AC', 'MU', 'PAL', 'ASM', 'FTGP',
        'FEL', 'FMA', 'FMT', 'FSC', 'LSP1', 'LSP2'}

    tail = [c for c in range(max(slots) + 1, ws.max_column + 1)]

    entries, duties = [], []
    for dayrow, day in dayrows:
        # Duty posts carry no times in this file: they sit in column 1 of the
        # block's third row, or in a spare column past the last period, and some
        # wrap over two rows ("BICYCLE" above "BAY"). Giving them a made-up slot
        # would put a teacher somewhere the school never said they were, so they
        # are kept as the day-level facts they are.
        posts, seen = [], set()
        for label in [g.get((dayrow + 1, 1), '')] + [g.get((r, c), '')
                                                     for r in (dayrow - 1, dayrow, dayrow + 1) for c in tail]:
            lab = tidy(label)
            if lab and lab.upper() not in DAYS and lab.upper() not in seen:
                seen.add(lab.upper())
                posts.append(lab)
        # "BUS STOP" + "GATE" is one post written over two lines, not two posts.
        joined = ' '.join(posts)
        for post in ([joined] if tidy(joined).upper() in DUTY else posts):
            duties.append({'day': day, 'post': tidy(post).title()})

        row = []
        for c in sorted(slots):
            got = resolve(g.get((dayrow - 1, c), ''), g.get((dayrow, c), ''), codes)
            row.append((c, got))
            if got and got[2] == 'lesson' and got[0].upper() not in codes:
                odd[got[0]] += 1

        # Adjacent columns holding the same thing are ONE block, not two.
        for c, got in row:
            if not got:
                continue
            title, klass, kind = got
            prev = entries[-1] if entries else None
            if (prev and prev['day'] == day and prev['title'] == title
                    and prev['class'] == klass and prev['end'] == slots[c][0]):
                prev['end'] = slots[c][1]
            else:
                entries.append({'cycle': None, 'day': day,
                                'start': slots[c][0], 'end': slots[c][1],
                                'title': title, 'class': klass, 'room': None, 'kind': kind})

    entries.sort(key=lambda e: (ORDER.index(e['day']), e['start']))
    # "TEACHER: Ms Ng Ai Leen (AILEEN)" → "Ms Ng Ai Leen". The parenthetical is
    # the sheet tag, which is what the sign-in handle is built from.
    full = re.sub(r'\s*\([^)]*\)\s*$', '', re.sub(r'^TEACHER:\s*', '', g.get((2, 1), ''), flags=re.I)).strip()
    if entries:
        teachers.append({'name': full or name.title(), 'sheet': name,
                         'entries': entries, 'dayDuties': duties})

print(json.dumps({'tag': tag, 'teachers': teachers,
                  'legendSubjects': legend_subj, 'legendRooms': legend_room,
                  'unknownTitles': dict(odd)}, indent=1))
