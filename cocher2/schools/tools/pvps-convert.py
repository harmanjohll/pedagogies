"""PVPS 2026 T3 staff timetable (.xlsx) → canonical Co-Cher entries.

One sheet per teacher. Each day column is read in 3-row blocks:
    row A: period number  | SUBJECT
    row B: start time     | CLASS          (or the whole label, for duty/assembly)
    row C: (blank)        | ROOM
Periods are 30 minutes from 0700. Consecutive identical blocks merge into one
entry, so a double period is one 60-minute block rather than two 30s.
"""
import openpyxl, json, re, sys, collections

SRC = sys.argv[1]
DAYS = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']

def hhmm(v):
    s = str(v).strip()
    m = re.fullmatch(r'(\d{1,2}):?(\d{2})', s)
    if m: return f'{int(m.group(1)):02d}:{m.group(2)}'
    m = re.fullmatch(r'(\d{3,4})', s)
    if m:
        t = m.group(1).zfill(4)
        return f'{t[:2]}:{t[2:]}'
    return None

def add30(t):
    h, m = map(int, t.split(':'))
    m += 30
    return f'{h + m // 60:02d}:{m % 60:02d}'

def kind_of(title):
    t = (title or '').upper()
    if t.startswith('DUTY') or t.startswith('UP@') or '@' in t: return 'duty'
    # Committees and councils are not lessons. Left as 'lesson' they both
    # mislabel the teacher's day and, worse, make "T&L Co" look like a subject
    # when departments are derived from what someone teaches.
    if re.search(r'\bPLT\b|\bSDC\b|\bSDT\b|\bPD\b|MEETING|MEET\b|STAFF ?MTG|DEPT|COUNCIL|T&L ?CO', t): return 'pd'
    if re.search(r'ASSEMBLY|FTGP|RECESS|CCA|FLAG|BREAK|LUNCH|VIA|VIRTUAL|PRAYER|REPORT LATE|IMP\b|ENRICHMENT', t): return 'school'
    return 'lesson'

def cell(ws, r, c):
    v = ws.cell(r, c).value
    return '' if v is None else str(v).strip().replace('\n', ' ')

wb = openpyxl.load_workbook(SRC, data_only=True)
teachers, legend_subj, legend_room, tag = [], {}, {}, ''

for name in wb.sheetnames:
    ws = wb[name]
    tag = tag or cell(ws, 1, 1)

    # Locate the header row, and the legend block that ends the timetable.
    hdr = next((r for r in range(1, ws.max_row + 1) if cell(ws, r, 1).upper().startswith('PERIOD')), None)
    if not hdr: continue
    stop = next((r for r in range(hdr + 1, ws.max_row + 1)
                 if cell(ws, r, 2).upper() in ('SUBJECTS', 'RESOURCES')), ws.max_row + 1)

    # Legend: short code → full name, for both subjects and rooms.
    for r in range(stop, ws.max_row + 1):
        for c in (2, 5):
            k, v = cell(ws, r, c), cell(ws, r, c + 1)
            if k and v and k.upper() not in ('SUBJECTS', 'RESOURCES'):
                (legend_room if re.match(r'^[A-Z]\d|^MEET|^HALL|ROOM|LAB', k.upper()) else legend_subj)[k] = v

    raw = collections.defaultdict(list)      # day → [(start, title, class, room)]
    r = hdr + 1
    while r < stop:
        start = hhmm(cell(ws, r + 1, 1)) if r + 1 < stop else None
        if not start: r += 1; continue
        for di, day in enumerate(DAYS):
            col = 2 + di
            subj, cls, room = cell(ws, r, col), cell(ws, r + 1, col), cell(ws, r + 2, col) if r + 2 <= ws.max_row else ''
            if not (subj or cls): continue
            # No subject on the period row means the middle cell IS the label —
            # that is how duties, assembly and FTGP are written in this file.
            title, klass = (subj, cls) if subj else (cls, '')
            if title == klass: klass = ''
            raw[day].append((start, title, klass, room))
        r += 3

    entries = []
    for day in DAYS:
        blocks = sorted(raw[day])
        for start, title, klass, room in blocks:
            prev = entries[-1] if entries else None
            same = (prev and prev['day'] == day and prev['title'] == title
                    and prev['class'] == (klass or None) and prev['room'] == (room or None)
                    and prev['end'] == start)
            if same:
                prev['end'] = add30(start)                  # extend the double period
            else:
                entries.append({'cycle': None, 'day': day, 'start': start, 'end': add30(start),
                                'title': title, 'class': klass or None, 'room': room or None,
                                'kind': kind_of(title)})
    if entries:
        teachers.append({'name': name.title(), 'sheet': name, 'entries': entries})

print(json.dumps({'tag': tag, 'teachers': teachers,
                  'legendSubjects': legend_subj, 'legendRooms': legend_room}, indent=1))
